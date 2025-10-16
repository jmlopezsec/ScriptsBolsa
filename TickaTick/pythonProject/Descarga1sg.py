'''
Este script descarga datos en barras en 1sg de Interactivebrokers.

Características principales del script:

1.- Un directorio por activo (ejemplo: ./data/AAPL/).
2.- Un fichero XLS por cada sesión de Wall Street (ejemplo: 2025-09-25.xls).
3.- La primera vez que se ejecuta para un activo:
    Crea el directorio si no existe.
    Intentará bajar el máximo histórico permitido
    Guardará los días en ficheros separados.
4.- Ejecuciones posteriores:
    Si estás en sesión (NY 9:30–16:00), solo baja los días anteriores que falten.
    Si estás fuera de sesión, baja el día en curso y los días anteriores faltantes.

se cambia del script DescargaTick.py la función fetch_ticks_for_session por fetch_bars_for_session

'''
#!/usr/bin/env python3
"""
Ajusta TICKERS, OUTPUT_ROOT, INIT_DAYS_BACK según necesites.
"""

import os
import time
from datetime import datetime, timedelta, time as dtime
import pytz
import pandas as pd
from ib_insync import IB, Stock, Future

import logging
from logging.handlers import TimedRotatingFileHandler



# ----------------------------
# CONFIGURACIÓN (TOP LEVEL)
# ----------------------------
TICKERS = ["SPY","PYPL","MO","KO","SBUX","AAPL","INTC","T"]    # Lista de símbolos
OUTPUT_ROOT = "E:/DATOSBOLSA/data1s"                # Carpeta raíz donde se crearán subdirectorios
#WHAT_TO_SHOW = "Bid_Ask"              # Usamos Bid_Ask para reqHistoricalTicks
WHAT_TO_SHOW = "Trades"              # Usamos Bid_Ask para reqHistoricalTicks
INIT_DAYS_BACK = 5                  # si no existe historial, intentamos éste nº de días atrás (ajustable)
IB_HOST = "127.0.0.1"
IB_PORT = 7496
IB_CLIENTID = 1

# Timezone / horario mercado
NY_TZ = pytz.timezone("America/New_York")
SESSION_OPEN = dtime(9, 30)
SESSION_CLOSE = dtime(16, 0)

#----------------------------------------
# Inicialización del Logging

# Configuración de logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)  # nivel mínimo que quieres capturar

# Formato de los mensajes
formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

# Handler para fichero rotativo (un archivo por día, guarda 7 días)
file_handler = TimedRotatingFileHandler(
    "ib_ticks.log", when="midnight", backupCount=7, encoding="utf-8"
)
file_handler.setFormatter(formatter)

# Handler para consola
console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)

# Añadir handlers al logger
logger.addHandler(file_handler)
logger.addHandler(console_handler)

#---------------------------------------


# ----------------------------
# UTILIDADES
# ----------------------------
def connect_ib():
    ib = IB()
    print(f"Conectando a IB en {IB_HOST}:{IB_PORT} (clientId={IB_CLIENTID})...", flush=True)
    logger.info(f"Conectando a IB en {IB_HOST}:{IB_PORT} (clientId={IB_CLIENTID})...")

    try:
        ib.connect(IB_HOST, IB_PORT, clientId=IB_CLIENTID, timeout=10)
        logger.info("** Conectado a IB")
    except Exception as e:
        print(f"❌ No se pudo conectar: {e}")
        logger.exception("❌ Error al conectar a IB")
        raise
    if not ib.isConnected():
        raise RuntimeError("❌ No se pudo establecer conexión con TWS/Gateway")
    print("✅ Conectado a IB")
    return ib

def ensure_symbol_dir(symbol):
    d = os.path.join(OUTPUT_ROOT, symbol)
    os.makedirs(d, exist_ok=True)
    return d

def list_downloaded_dates(symbol_dir):
    """Devuelve set de fechas descargadas en formato date (YYYY-MM-DD.xlsx)"""
    if not os.path.exists(symbol_dir):
        return set()
    names = os.listdir(symbol_dir)
    dates = set()
    for n in names:
        # aceptar .xlsx o .xls
        if n.lower().endswith(".xlsx") or n.lower().endswith(".xls"):
            base = os.path.splitext(n)[0]
            try:
                dt = datetime.strptime(base, "%Y-%m-%d").date()
                dates.add(dt)
            except Exception:
                continue
    return dates

def is_market_open(now=None):
    """True si ahora (NY) está entre 9:30 y 16:00 (sesión regular)."""
    now = now or datetime.now(pytz.utc).astimezone(NY_TZ)
    tod = now.time()
    return (tod >= SESSION_OPEN) and (tod <= SESSION_CLOSE) and (now.weekday() < 5)

def next_business_day(date):
    d = date + timedelta(days=1)
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d

def prev_business_day(date):
    d = date - timedelta(days=1)
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d

# Helper robusto para extraer campos de ticks (bid/ask)
def _get_attr(obj, names):
    for n in names:
        if hasattr(obj, n):
            return getattr(obj, n)
    return None

def tick_obj_to_row(tick):
    """
    Convierte un objeto HistoricalTickBidAsk (o similar) a dict con:
    time (tz-aware UTC), bid, bid_size, ask, ask_size, midpoint
    """
    # time puede venir tz-aware o naive; convertir/asegurar UTC tz-aware
    ttime = getattr(tick, "time", None)
    if ttime is None:
        return None

    #Normalizar time a UTC
    # si naive, interpretamos como UTC
    if ttime.tzinfo is None:
       ttime = pytz.utc.localize(ttime)
    else:
       # normalizar a UTC
       ttime = ttime.astimezone(pytz.utc)


    if WHAT_TO_SHOW == "Bid_Ask":
        bid = _get_attr(tick, ["priceBid", "bidPrice", "bid"])
        ask = _get_attr(tick, ["priceAsk", "askPrice", "ask"])
        bsize = _get_attr(tick, ["sizeBid", "bidSize", "size_bid"])
        asize = _get_attr(tick, ["sizeAsk", "askSize", "size_ask"])

        # forzar floats o None
        def _flt(x):
            try:
                return float(x) if x is not None else None
            except Exception:
                return None

        bid, ask, bsize, asize = map(_flt, (bid, ask, bsize, asize))
        midpoint = (bid+ask) / 2.0 if (bid and ask) else None

        return {
            "time": ttime,
            "bid": bid,
            "bid_size": bsize,
            "ask": ask,
            "ask_size": asize,
            "midpoint": midpoint
        }
    else: #WHAT_TO_SHOW == "last"
        price = _get_attr(tick, ["price", "lastPrice"])
        size = _get_attr(tick, ["size", "lastSize"])
        try:
            price = float(price)
        except Exception:
            price = None
        try:
            size = float(size)
        except Exception:
            size = None
        return {
            "time": ttime,
            "price": price,
            "size": size
        }


# ----------------------------
# DESCARGA ENCadenada para UN DÍA
# ----------------------------
'''
def fetch_ticks_for_session(ib, contract, session_start_ny, session_end_ny):
    """
    Descarga todos los ticks de la sesión [session_start_ny, session_end_ny] (ambos tz-aware NY)
    encadenando llamadas de 1000 en 1000 hacia atrás hasta cubrir la sesión completa.
    Devuelve DataFrame con columnas: time (UTC tz-aware), bid, bid_size, ask, ask_size, midpoint

    Descarga todos los ticks de la sesión encadenando bloques de 1000 ticks hacia adelante.
    Soporta Bid_Ask y Last según WHAT_TO_SHOW


    """
    # convertir a UTC tz-aware datetimes
    start_utc = session_start_ny.astimezone(pytz.utc)
    end_utc = session_end_ny.astimezone(pytz.utc)

    key= None
    rows = []
    seen_keys = set()  # para evitar duplicados por solapamientos entre pages

    #Antes current_start = start_utc
    current_end= end_utc

    iteration = 0
    max_time = None

    while current_end >= start_utc:
        iteration += 1
        # IB acepta endDateTime como string 'YYYYMMDD HH:MM:SS' (UTC)
        #Antes start_str = current_start.strftime("%Y%m%d %H:%M:%S")
        #Antes end_str = end_utc.strftime("%Y%m%d %H:%M:%S")

        end_str = current_end.strftime("%Y%m%d %H:%M:%S")

        logger.debug(f"Iteración {iteration} tiempo a pedir: {end_str}")

        try:
            ticks = ib.reqHistoricalTicks(
                contract,
                #startDateTime=start_str,                   # dejamos vacío para pedir max hacia atrás desde endDateTime
                startDateTime="",
                endDateTime=current_end, # end_str,
                numberOfTicks=1000,
                whatToShow=WHAT_TO_SHOW,
                useRth=True,
                ignoreSize=False
            )
            logger.debug(f"Recibidos {len(ticks)} ticks en iter {iteration} ({start_utc} → {end_str})")
        except Exception as e:
            print(f"    ⚠️ reqHistoricalTicks fallo en iter {iteration} hasta {end_str}: {e}")
            logger.exception(f"⚠️ reqHistoricalTicks fallo en iter {iteration} hasta {end_str}")
            # si falla por pacing, esperar y reintentar
            time.sleep(1.0)
            continue

        if not ticks:
            break

        first_tick = ticks[0]
        last_tick = ticks[-1]
        logger.debug(
            f"Primer tick: time={first_tick.time}, price={getattr(first_tick, 'price', None)}, bid={getattr(first_tick, 'priceBid', None)}, ask={getattr(first_tick, 'priceAsk', None)}")
        logger.debug(
            f"Último tick: time={last_tick.time}, price={getattr(last_tick, 'price', None)}, bid={getattr(last_tick, 'priceBid', None)}, ask={getattr(last_tick, 'priceAsk', None)}")


        # convertir y filtrar por >0 start-utc Usar dedupe por key compuesto para evitar duplicados entre páginas

        block_rows = []
        block_times = []

        for t in ticks:
            row = tick_obj_to_row(t)
            if row is None:
                continue

            ttime = row["time"] #tz-aware UTC
            logger.debug (f"tiempo tick: {ttime}")
            # filtrar fuera de sesión (anterior al inicio)
            if ttime < start_utc or ttime > end_utc:
                logger.info("Fuera de sesión")
                continue

            if "bid" in row:  # estamos en Bid_Ask
                key = (
                    ttime.isoformat(),
                    row.get("bid"),
                    row.get("ask"),
                    row.get("bid_size"),
                    row.get("ask_size")
                )
            elif "price" in row:  # estamos en Trades/Last
                key = (
                    ttime.isoformat(),
                    row.get("price"),
                    row.get("size")
                )
            else:  # fallback genérico
                key = tuple(row.items())



            if key in seen_keys:
                continue

            logger.debug(f"Key generada: {key}")
            seen_keys.add(key)
            block_rows.append(row)
            block_times.append(ttime)
            #logger.debug(f"Append block_times: {ttime}")

        if block_rows:
            rows.extend(block_rows)

        if not block_rows:
            break

        # retroceder el end actual justo antes del tick más antiguo del bloque
        earliest = min(block_times)
        logger.debug(f"El earliest: {earliest}")
        current_end = earliest - timedelta(seconds=1)  # IB solo entiende segundos

        time.sleep(0.2)

    logger.debug(f"valor current_end: {current_end}")
    # al final, ordenar todos los ticks ascendentemente por tiempo
    if rows:
        df = pd.DataFrame(rows)
        df.sort_values("time", inplace=True)
        df["time"] = pd.to_datetime(df["time"]).dt.tz_convert(pytz.utc)
    else:
        df = pd.DataFrame(columns=["time", "bid", "bid_size", "ask", "ask_size", "midpoint"])

    return df
'''

# ----------------------------
# DESCARGA DE BARRAS DE 1 SEGUNDO PARA UNA SESIÓN
# ----------------------------


# ----------------------------
# DESCARGA DE BARRAS 1 SEGUNDO
# ----------------------------
def fetch_ticks_for_session(ib, contract, session_start_ny, session_end_ny):
    """
    Descarga todos los datos de 1 segundo de la sesión [session_start_ny, session_end_ny],
    troceando en bloques de 30 minutos (IB solo permite 1s con duraciones <= 1h).
    Devuelve DataFrame con columnas: time, open, high, low, close, volume.
    """
    start_utc = session_start_ny.astimezone(pytz.utc)
    end_utc = session_end_ny.astimezone(pytz.utc)

    rows = []
    block_start = start_utc
    iteration = 0

    while block_start < end_utc:
        iteration += 1
        block_end = min(block_start + timedelta(minutes=30), end_utc)
        logger.debug(f"Iteración {iteration}: {block_start} → {block_end}")

        try:
            bars = ib.reqHistoricalData(
                contract,
                endDateTime=block_end,
                durationStr="1800 S",       # 30 minutos = 1800 segundos
                barSizeSetting="1 secs",    # barras de 1 segundo
                whatToShow=WHAT_TO_SHOW,
                useRTH=True,
                formatDate=1,
                keepUpToDate=False
            )
            logger.debug(f"Recibidas {len(bars)} barras de 1s en iter {iteration}")
        except Exception as e:
            logger.error(f"⚠️ reqHistoricalData fallo en iter {iteration}: {e}")
            time.sleep(1.0)
            continue

        if not bars:
            logger.info(f"Sin datos en iter {iteration}")
            block_start = block_end
            continue

        for b in bars:
            ttime = b.date
            if ttime.tzinfo is None:
                ttime = pytz.utc.localize(ttime)
            else:
                ttime = ttime.astimezone(pytz.utc)

            if ttime < start_utc or ttime > end_utc:
                continue

            rows.append({
                "time": ttime,
                "open": b.open,
                "high": b.high,
                "low": b.low,
                "close": b.close,
                "volume": b.volume
            })

        block_start = block_end
        time.sleep(1)

    if rows:
        df = pd.DataFrame(rows)
        df.sort_values("time", inplace=True)
    else:
        df = pd.DataFrame(columns=["time", "open", "high", "low", "close", "volume"])

    return df



# ----------------------------
# GUARDADO por DÍA (XLSX)
# ----------------------------
def save_session_df(symbol_dir, date_obj, df):
    """
    Guarda dataframe de sesión en symbol_dir/YYYY-MM-DD.xlsx
    Convierte time a tz-naive UTC antes de escribir y crea formato de columna.
    """
    from openpyxl import load_workbook

    filename = os.path.join(symbol_dir, f"{date_obj.strftime('%Y-%m-%d')}.xlsx")
    if df.empty:
        print(f"    (vacío) No se guarda {filename}")
        return

    # convertir time tz-aware UTC -> tz-naive (Excel no acepta tz-aware)
    df_to_write = df.copy()
    df_to_write["time"] = pd.to_datetime(df_to_write["time"]).dt.tz_convert(pytz.utc).dt.tz_localize(None)

    # Guardar con pandas (openpyxl) y luego aplicar formato a columna A
    df_to_write.to_excel(filename, index=False)

    # Formatear la columna A como datetime 'yyyy-mm-dd hh:mm:ss' usando openpyxl
    try:
        wb = load_workbook(filename)
        ws = wb.active
        for cell in ws["A"][2:]:  # saltar encabezado (fila 1)
            cell.number_format = "yyyy-mm-dd hh:mm:ss"
        wb.save(filename)
    except Exception as e:
        print(f"    ⚠️ No se pudo aplicar formato a {filename}: {e}")

    print(f"    💾 Guardado {filename} ({len(df)} filas)")

# ----------------------------
# LÓGICA PRINCIPAL
# ----------------------------
def main():
    ib = connect_ib()

    today_ny = datetime.now(pytz.utc).astimezone(NY_TZ).date()
    market_open_now = is_market_open()

    for symbol in TICKERS:
        print(f"\n== Procesando {symbol} ==")
        logger.info(f"== Procesando {symbol} ==")
        symbol_dir = ensure_symbol_dir(symbol)
        downloaded = list_downloaded_dates(symbol_dir)
        # definir rango a descargar

        if not downloaded:
            # primer arranque: intentamos INIT_DAYS_BACK días atrás (ajustando fines de semana)
            end_date = today_ny
            if market_open_now:
                end_date = prev_business_day(today_ny)  # no descargar sesión en curso
            # retroceder hasta obtener INIT_DAYS_BACK business days
            start_date = end_date
            days_added = 0
            while days_added < INIT_DAYS_BACK:
                if start_date.weekday() < 5:
                    days_added += 1
                start_date -= timedelta(days=1)
            # after loop start_date is 1 day earlier than needed
            start_date = next_business_day(start_date)
        else:
            last = max(downloaded)
            start_date = next_business_day(last)
            if market_open_now:
                end_date = prev_business_day(today_ny)
            else:
                end_date = today_ny

        if start_date > end_date:
            print("  ✅ No hay días nuevos que descargar")
            continue

        print(f"  Descargando rango: {start_date} → {end_date} (market_open_now={market_open_now})")
        logger.debug(f"Descargando rango: {start_date} → {end_date} (market_open_now={market_open_now})")

        # preparar contrato
        contract = Stock(symbol, "SMART", "USD")
        # No funciona contract = Future(symbol, '202509', 'GLOBEX')
        try:
            ib.qualifyContracts(contract)
        except Exception as e:
            print(f"  ⚠️ qualifyContracts fallo para {symbol}: {e}")

        d = start_date
        while d <= end_date:
            if d.weekday() >= 5:
                d = next_business_day(d)
                continue
            if d in downloaded:
                print(f"    {d} ya descargado → saltando")
                d = next_business_day(d)
                continue

            print(f"    → Descargando sesión {d}")
            session_start = NY_TZ.localize(datetime.combine(d, SESSION_OPEN))
            session_end = NY_TZ.localize(datetime.combine(d, SESSION_CLOSE))
            try:
                df_day = fetch_ticks_for_session(ib, contract, session_start, session_end)
                if not df_day.empty:
                    save_session_df(symbol_dir, d, df_day)
                    time.sleep(60.0) #Evitar el pacing
                else:
                    print(f"    ⚠️ No hubo ticks para {symbol} {d}")
            except Exception as e:
                print(f"    ❌ Error Descargar {symbol} {d}: {e}")

            d = next_business_day(d)

    ib.disconnect()
    print("\n✂ Desconectado. Proceso finalizado.")

if __name__ == "__main__":
    main()
